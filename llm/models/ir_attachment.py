# -*- coding: utf-8 -*-
"""Extracción de texto de adjuntos para el contexto LLM."""

import base64
import io
import logging

from odoo import models
from odoo.tools import html2plaintext

_logger = logging.getLogger(__name__)

# Límite por archivo para no saturar el contexto del modelo
MAX_LLM_TEXT_CHARS = 200000


class IrAttachment(models.Model):
    _inherit = "ir.attachment"

    def _llm_get_raw_bytes(self):
        self.ensure_one()
        if self.raw:
            return self.raw
        if self.datas:
            if isinstance(self.datas, bytes):
                return self.datas
            return base64.b64decode(self.datas)
        return b""

    def llm_extract_text(self):
        """Devuelve texto plano extraído del adjunto para enviarlo al modelo, o cadena vacía."""
        self.ensure_one()
        raw = self._llm_get_raw_bytes()
        if not raw:
            return ""
        mimetype = (self.mimetype or "").lower()
        name = (self.name or "").lower()

        try:
            if mimetype == "text/html" or name.endswith((".html", ".htm")):
                return html2plaintext(
                    raw.decode("utf-8", errors="replace")
                )[:MAX_LLM_TEXT_CHARS]

            if mimetype.startswith("text/") or mimetype == "application/json":
                return raw.decode("utf-8", errors="replace")[:MAX_LLM_TEXT_CHARS]

            # PDF
            if mimetype == "application/pdf" or name.endswith(".pdf"):
                return self._llm_extract_pdf_text(raw)

            # Word .docx
            if "wordprocessingml" in mimetype or name.endswith(".docx"):
                return self._llm_extract_docx_text(raw)

            # Excel: .xlsx/.xlsm (openpyxl); .xls antiguo (xlrd). Ojo: str.endswith(".xls") coincide con ".xlsx".
            xlsx_ext = name.endswith((".xlsx", ".xlsm"))
            xls_legacy_ext = name.endswith(".xls") and not xlsx_ext
            if xlsx_ext or "spreadsheetml" in mimetype:
                return self._llm_extract_xlsx_text(raw)
            if xls_legacy_ext or mimetype in (
                "application/vnd.ms-excel",
                "application/excel",
            ):
                return self._llm_extract_xls_text(raw)

        except Exception as err:
            _logger.warning(
                "No se pudo extraer texto del adjunto %s (%s): %s",
                self.id,
                self.name,
                err,
            )
        return ""

    def _llm_extract_pdf_text(self, raw_bytes):
        try:
            from pypdf import PdfReader
        except ImportError:
            try:
                from PyPDF2 import PdfReader
            except ImportError:
                _logger.warning(
                    "Instale el paquete Python 'pypdf' para extraer texto de PDFs en el chat LLM."
                )
                return ""
        try:
            reader = PdfReader(io.BytesIO(raw_bytes))
            chunks = []
            for page in reader.pages:
                chunks.append(page.extract_text() or "")
            text = "\n".join(chunks).strip()
            return text[:MAX_LLM_TEXT_CHARS]
        except Exception as err:
            _logger.warning("Error leyendo PDF adjunto: %s", err)
            return ""

    def _llm_extract_docx_text(self, raw_bytes):
        try:
            import docx
        except ImportError:
            _logger.debug("python-docx no instalado; omitiendo .docx")
            return ""
        try:
            document = docx.Document(io.BytesIO(raw_bytes))
            return "\n".join(p.text for p in document.paragraphs)[:MAX_LLM_TEXT_CHARS]
        except Exception as err:
            _logger.warning("Error leyendo DOCX: %s", err)
            return ""

    def _llm_extract_xlsx_text(self, raw_bytes):
        """Excel moderno (.xlsx, .xlsm): tablas como texto tabulado por hoja."""
        try:
            import openpyxl
        except ImportError:
            _logger.warning(
                "Instale openpyxl para extraer texto de Excel en el chat LLM: pip install openpyxl"
            )
            return ""
        max_rows_per_sheet = 8000
        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(raw_bytes), read_only=True, data_only=True
            )
            parts = []
            try:
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    lines = []
                    for i, row in enumerate(sheet.iter_rows(values_only=True)):
                        if i >= max_rows_per_sheet:
                            lines.append(
                                f"... [filas omitidas: más de {max_rows_per_sheet} en esta hoja] ..."
                            )
                            break
                        cells = []
                        for c in row:
                            cells.append("" if c is None else str(c).strip())
                        lines.append("\t".join(cells))
                    parts.append(
                        f"=== Hoja: {sheet_name} ===\n" + "\n".join(lines)
                    )
            finally:
                wb.close()
            text = "\n\n".join(parts).strip()
            return text[:MAX_LLM_TEXT_CHARS]
        except Exception as err:
            _logger.warning("Error leyendo Excel (.xlsx): %s", err)
            return ""

    def _llm_extract_xls_text(self, raw_bytes):
        """Excel binario antiguo (.xls) con xlrd, si está disponible."""
        try:
            import xlrd
        except ImportError:
            _logger.warning(
                "Para leer archivos .xls antiguos instale xlrd: pip install xlrd"
            )
            return ""
        max_rows_per_sheet = 8000
        try:
            book = xlrd.open_workbook(file_contents=raw_bytes)
            parts = []
            for sheet_ix in range(book.nsheets):
                sh = book.sheet_by_index(sheet_ix)
                lines = []
                nrows = min(sh.nrows, max_rows_per_sheet)
                for rx in range(nrows):
                    row = sh.row(rx)
                    lines.append(
                        "\t".join(
                            str(c.value) if c.value != "" else ""
                            for c in row
                        )
                    )
                if sh.nrows > max_rows_per_sheet:
                    lines.append(
                        f"... [filas omitidas: más de {max_rows_per_sheet} en esta hoja] ..."
                    )
                parts.append(f"=== Hoja: {sh.name} ===\n" + "\n".join(lines))
            text = "\n\n".join(parts).strip()
            return text[:MAX_LLM_TEXT_CHARS]
        except Exception as err:
            _logger.warning("Error leyendo Excel (.xls): %s", err)
            return ""
